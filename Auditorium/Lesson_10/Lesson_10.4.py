import openpyxl
wb = openpyxl.load_workbook("../../Texts/text_05.xlsx")
print(wb.sheetnames)
ws = wb.active
ws["A1"].value = 100
ws["B2"].value = 200
ws["C3"].value = ws["A1"].value + ws["B2"].value
ws["C3"].value
ws.cell(3, 3).value
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        print(i, j, ws.cell(i, j).value)