import openpyxl
wb = openpyxl.load_workbook("../../Texts/text_05.xlsx")
ws = wb["Лист1"]
su = 0
for i in range(1, ws.max_row + 1):
    su += ws.cell(i, 2).value
j = ws.max_row + 1
ws.cell(j, 1).value = "ИТОГО"
ws.cell(j, 2).value = su
wb.save("../Texts/text_05.xlsx")