import openpyxl
wb = openpyxl.load_workbook("../../Texts/text_07.xlsx")
for sh in wb.sheetnames:
    ws = wb[sh]
    print(sh, "------------")
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            c = ws.cell(i, j)
            if c.value != None:
                print(c.coordinate, c.value, end = "/")
print()