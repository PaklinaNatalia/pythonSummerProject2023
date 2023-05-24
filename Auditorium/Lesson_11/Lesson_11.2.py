import openpyxl
wb = openpyxl.load_workbook("../../Texts/text_05.xlsx")
ws = wb["Лист1"]
for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        c = ws.cell(i, j)
        if c.value != None:
            print(c.coordinate, c.value)