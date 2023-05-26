import openpyxl
wb = openpyxl.load_workbook("../../Texts/text_06.xlsx")
ws = wb.active
d = {}
for i in range(1, ws.max_row + 1):
    name = ws.cell (i, 1).value
    val = int(ws.cell(i, 2).value)
    d[name] = d.get(name, 0) + val
d["Итого"] = d.get("Итого", sum(d.values()))
wb.create_sheet("Сводная таблица")
wb.active = wb["Сводная таблица"]
ws = wb.active
i = 1
for k, v in d.items():
    ws.cell(i, 1).value = k
    ws.cell(i, 2).value = v
    i += 1
print("Готово!")
wb.save("../Texts/text_06.xlsx")