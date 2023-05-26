import openpyxl, csv
from openpyxl import Workbook

wb = Workbook()
wb.save("../Texts/staff.xlsx")
wb = openpyxl.load_workbook("../../Texts/staff.xlsx")
wb.create_sheet("Лист1")
ws = wb["Лист1"]

with open("../../Texts/staff.csv") as file:
    rows = csv.DictReader(file)
    su = 0
    lst = []
    for row in rows:
        lst.append(row)
    for i in lst[1:]:
        su += int(i[4])
    lst0 = lst[0]
    new_lst = sorted(lst[1:], key = lambda x: (x[3], x[1], x[2]))
    print(lst0, new_lst, su)

ws.append(lst0)
for k in new_lst:
    ws.append(k)
ws.cell(ws.max_row + 1, 1).value = "Итого"
ws.cell(ws.max_row, 5).value = su
wb.save("../Texts/staff.xlsx")